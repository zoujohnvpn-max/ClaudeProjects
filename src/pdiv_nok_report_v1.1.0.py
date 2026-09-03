# -*- coding: utf-8 -*-
# =============================================
# Author: John
# Date: 2026-09-03
# Description: Split PDIV NOK raw data into station A/B sheets, compute statistics and build line charts
# =============================================
# --- Version History ---
# v1.0.0 (2026-09-03): initial release, split A/B stations, add summary sheet and line charts
# v1.1.0 (2026-09-03): add dedicated chart sheet with per-metric line charts, fix category axis
#                      position, use string references for date categories, apply validated palette
# =============================================

VERSION = "v1.1.0"

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------- 样式常量 ----------------
FONT_TITLE = Font(name="微软雅黑", size=13, bold=True, color="1F3864")
FONT_SUB = Font(name="微软雅黑", size=11, bold=True, color="52514E")
FONT_HEAD = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="微软雅黑", size=10)
FONT_TOTAL = Font(name="微软雅黑", size=10, bold=True, color="1F3864")
FONT_NOTE = Font(name="微软雅黑", size=9, color="595959")

FILL_A = PatternFill("solid", fgColor="184F95")   # A 站表头（蓝）
FILL_B = PatternFill("solid", fgColor="A3400F")   # B 站表头（橙）
FILL_SUM = PatternFill("solid", fgColor="375623")  # 汇总表头（绿）
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# 配色（经 dataviz validate_palette 校验）：
# A 站 = 蓝色序数梯度，B 站 = 橙色序数梯度，PDIV 次数越靠前颜色越深；
# A/B 对比图用蓝 / 橙两个分类色，CVD 与常规视觉分离度均达标。
RAMP_A = ["184F95", "3987E5", "86B6EF"]
RAMP_B = ["A3400F", "EB6834", "F0906A"]
COLOR_A, COLOR_B = "2A78D6", "EB6834"

SHEET_A, SHEET_B = "A站明细", "B站明细"

HEADERS = [
    "日期", "生产数量",
    "PDIV 1次 NOK", "PDIV 1次 不合格率",
    "PDIV 2次 NOK", "PDIV 2次 不合格率",
    "PDIV 3次 NOK", "PDIV 3次 不合格率",
    "NOK 合计", "合计不合格率",
]


# ---------------- 原始数据解析 ----------------
def parse_nok(value):
    """把 '3/1.4%'、'0'、'2/0.95' 这类文本解析成 NOK 数量（整数）。"""
    if value is None:
        return 0
    head = str(value).strip().split("/")[0].strip()
    m = re.search(r"-?\d+(?:\.\d+)?", head)
    return int(float(m.group())) if m else 0


def parse_int(value):
    if value is None:
        return 0
    m = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return int(float(m.group())) if m else 0


def read_raw(path):
    """读取原始表，返回 [(日期, A站产量, A1, A2, A3, B站产量, B1, B2, B3), ...]。"""
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=4, max_col=9, values_only=True):
        date = row[0]
        if date is None or not str(date).strip():
            continue
        rows.append((
            str(date).strip(),
            parse_int(row[1]), parse_nok(row[2]), parse_nok(row[3]), parse_nok(row[4]),
            parse_int(row[5]), parse_nok(row[6]), parse_nok(row[7]), parse_nok(row[8]),
        ))
    return rows


# ---------------- 图表工厂 ----------------
def make_line_chart(title, y_title, num_fmt, series_defs, cat_sheet, cat_range,
                    width=26, height=9):
    """按统一规范生成折线图。

    series_defs: [(数据源 worksheet, 列号, 线条颜色), ...]，标题取该列第 2 行表头。
    cat_range:   分类轴（日期）区间，形如 "$A$3:$A$33"。
    """
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height, chart.width = height, width
    chart.y_axis.title = y_title
    chart.y_axis.numFmt = num_fmt
    chart.x_axis.title = "日期"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.axPos = "b"          # 分类轴必须在下方，openpyxl 默认值是 "l"
    chart.y_axis.axPos = "l"
    chart.x_axis.tickLblSkip = 2      # 31 个日期，隔一个显示，避免标签重叠
    chart.x_axis.tickMarkSkip = 2

    for ws, col, color in series_defs:
        first, last = cat_range.split(":")
        ref = Reference(ws, min_col=col, min_row=2,
                        max_row=int(re.search(r"\d+$", last).group()))
        chart.add_data(ref, titles_from_data=True)

    # 日期是文本，用 StrRef 而不是 openpyxl 默认的 NumRef，否则部分阅读器把日期当数字
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f="'{}'!{}".format(cat_sheet, cat_range)))

    for series, (_, _, color) in zip(chart.series, series_defs):
        series.smooth = False
        series.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
        series.marker = Marker(symbol="circle", size=6)
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color

    # 单条曲线不需要图例，标题已说明是哪条；两条以上必须有图例
    if len(series_defs) == 1:
        chart.legend = None
    else:
        chart.legend.position = "b"
        chart.legend.overlay = False
    return chart


# ---------------- 明细表 ----------------
def style_table(ws, head_fill, n_rows, n_cols=10):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, head_fill, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    for r in range(3, 3 + n_rows):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col in (4, 6, 8, 10):
                cell.number_format = "0.00%"

    total_row = 3 + n_rows
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_TOTAL, FILL_TOTAL, CENTER, BORDER
        if col in (4, 6, 8, 10):
            cell.number_format = "0.00%"

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 11
    for col in "CEGI":
        ws.column_dimensions[col].width = 15
    for col in "DFHJ":
        ws.column_dimensions[col].width = 17
    ws.freeze_panes = "B3"


def write_station_sheet(wb, title, station_name, data, head_fill, ramp):
    """单个工位的明细表，表格下方附该工位的三条 PDIV 不合格率曲线。"""
    ws = wb.create_sheet(title)
    ws["A1"] = f"{station_name} PDIV 检测数据明细（{data[0][0]} ~ {data[-1][0]}）"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.row_dimensions[1].height = 24

    for col, head in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col, value=head)

    for i, (date, qty, n1, n2, n3) in enumerate(data):
        r = 3 + i
        total_nok = n1 + n2 + n3
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=qty)
        for j, nok in enumerate((n1, n2, n3)):
            ws.cell(row=r, column=3 + j * 2, value=nok)
            ws.cell(row=r, column=4 + j * 2, value=(nok / qty if qty else 0))
        ws.cell(row=r, column=9, value=total_nok)
        ws.cell(row=r, column=10, value=(total_nok / qty if qty else 0))

    n = len(data)
    total_row = 3 + n
    ws.cell(row=total_row, column=1, value="合计 / 平均")
    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row - 1})")
    for col in (3, 5, 7, 9):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}3:{letter}{total_row - 1})")
        ws.cell(row=total_row, column=col + 1,
                value=f"=IF($B${total_row}=0,0,{letter}{total_row}/$B${total_row})")

    style_table(ws, head_fill, n)

    cat_range = f"$A$3:$A${total_row - 1}"
    ws.cell(row=total_row + 2, column=1, value=f"▼ {station_name} PDIV 不合格率趋势").font = FONT_SUB
    chart = make_line_chart(
        f"{station_name} PDIV 不合格率趋势", "不合格率", "0.0%",
        [(ws, 4, ramp[0]), (ws, 6, ramp[1]), (ws, 8, ramp[2])],
        title, cat_range,
    )
    ws.add_chart(chart, f"A{total_row + 3}")
    return ws


# ---------------- 图表专页 ----------------
def write_chart_sheet(wb, ws_a, ws_b, n_days):
    """所有曲线图集中放在这张表，打开工作簿第一眼就能看到。"""
    ws = wb.create_sheet("趋势图表")
    last = 2 + n_days
    cat = f"$A$3:$A${last}"

    ws["A1"] = "PDIV 检测不合格率趋势图"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "数据源：A站明细 / B站明细 两张表，改动明细数据后图表自动更新。"
    ws["A2"].font = FONT_NOTE
    ws.column_dimensions["A"].width = 4

    specs = [
        ("A 站 · PDIV 1次 不合格率", [(ws_a, 4, COLOR_A)], SHEET_A, "不合格率", "0.0%"),
        ("B 站 · PDIV 1次 不合格率", [(ws_b, 4, COLOR_B)], SHEET_B, "不合格率", "0.0%"),
        ("A 站 · PDIV 1次 / 2次 / 3次 不合格率",
         [(ws_a, 4, RAMP_A[0]), (ws_a, 6, RAMP_A[1]), (ws_a, 8, RAMP_A[2])], SHEET_A, "不合格率", "0.0%"),
        ("B 站 · PDIV 1次 / 2次 / 3次 不合格率",
         [(ws_b, 4, RAMP_B[0]), (ws_b, 6, RAMP_B[1]), (ws_b, 8, RAMP_B[2])], SHEET_B, "不合格率", "0.0%"),
        ("A 站 · 合计不合格率", [(ws_a, 10, COLOR_A)], SHEET_A, "不合格率", "0.0%"),
        ("B 站 · 合计不合格率", [(ws_b, 10, COLOR_B)], SHEET_B, "不合格率", "0.0%"),
        ("A 站 · 每日生产数量", [(ws_a, 2, COLOR_A)], SHEET_A, "台数", "General"),
        ("B 站 · 每日生产数量", [(ws_b, 2, COLOR_B)], SHEET_B, "台数", "General"),
    ]

    row = 4
    for title, series_defs, cat_sheet, y_title, fmt in specs:
        ws.cell(row=row, column=1, value=f"▼ {title}").font = FONT_SUB
        chart = make_line_chart(title, y_title, fmt, series_defs, cat_sheet, cat)
        ws.add_chart(chart, f"A{row + 1}")
        row += 21
    return ws


def write_compare_sheet(wb, data_a, data_b, dates):
    """A / B 站逐日对比明细 + 对比曲线。"""
    ws = wb.create_sheet("A_B站对比")
    ws["A1"] = "A 站 / B 站 逐日对比"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 24

    heads = ["日期", "A站 产量", "B站 产量",
             "A站 PDIV1 不合格率", "B站 PDIV1 不合格率",
             "A站 PDIV2 不合格率", "B站 PDIV2 不合格率",
             "A站 合计不合格率", "B站 合计不合格率"]
    for col, head in enumerate(heads, start=1):
        cell = ws.cell(row=2, column=col, value=head)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, FILL_SUM, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    def rate(rec, idx):
        return (rec[idx] / rec[0]) if rec[0] else 0

    for i, date in enumerate(dates):
        r = 3 + i
        a, b = data_a[i][1:], data_b[i][1:]
        values = [date, a[0], b[0],
                  rate(a, 1), rate(b, 1), rate(a, 2), rate(b, 2),
                  (sum(a[1:]) / a[0] if a[0] else 0), (sum(b[1:]) / b[0] if b[0] else 0)]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col >= 4:
                cell.number_format = "0.00%"

    last = 2 + len(dates)
    ws.column_dimensions["A"].width = 13
    for col in "BC":
        ws.column_dimensions[col].width = 12
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 20
    ws.freeze_panes = "B3"

    cat = f"$A$3:$A${last}"
    charts = [
        ("PDIV 1次 不合格率对比（A 站 vs B 站）", [(ws, 4, COLOR_A), (ws, 5, COLOR_B)], "不合格率", "0.0%"),
        ("PDIV 2次 不合格率对比（A 站 vs B 站）", [(ws, 6, COLOR_A), (ws, 7, COLOR_B)], "不合格率", "0.0%"),
        ("合计不合格率对比（PDIV 1+2+3次）", [(ws, 8, COLOR_A), (ws, 9, COLOR_B)], "不合格率", "0.0%"),
        ("每日产量对比（A 站 vs B 站）", [(ws, 2, COLOR_A), (ws, 3, COLOR_B)], "台数", "General"),
    ]
    row = last + 2
    for title, series_defs, y_title, fmt in charts:
        ws.cell(row=row, column=1, value=f"▼ {title}").font = FONT_SUB
        ws.add_chart(make_line_chart(title, y_title, fmt, series_defs, "A_B站对比", cat), f"A{row + 1}")
        row += 21
    return ws


# ---------------- 统计汇总 ----------------
def station_stats(records):
    """records: [(qty, n1, n2, n3), ...] -> 统计字典。"""
    qty = sum(r[0] for r in records)
    noks = [sum(r[i] for r in records) for i in (1, 2, 3)]
    total_nok = sum(noks)
    daily = [(sum(r[1:]) / r[0]) if r[0] else 0 for r in records]
    return {
        "days": len(records), "qty": qty, "noks": noks, "total_nok": total_nok,
        "rates": [n / qty if qty else 0 for n in noks],
        "total_rate": total_nok / qty if qty else 0,
        "max_rate": max(daily), "max_idx": daily.index(max(daily)),
        "min_rate": min(daily), "min_idx": daily.index(min(daily)),
        "avg_qty": qty / len(records) if records else 0,
    }


def write_summary_sheet(wb, dates, data_a, data_b):
    ws = wb.create_sheet("统计汇总")
    ws["A1"] = f"PDIV 检测数据统计汇总（{dates[0]} ~ {dates[-1]}，共 {len(dates)} 个生产日）"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 26

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])

    rows = [
        ("统计项", "A 站", "B 站", "差值（B - A）", None),
        ("生产日数（天）", sa["days"], sb["days"], sb["days"] - sa["days"], "int"),
        ("总生产数量（台）", sa["qty"], sb["qty"], sb["qty"] - sa["qty"], "int"),
        ("日均生产数量（台）", sa["avg_qty"], sb["avg_qty"], sb["avg_qty"] - sa["avg_qty"], "f1"),
        ("PDIV 1次 NOK 总数", sa["noks"][0], sb["noks"][0], sb["noks"][0] - sa["noks"][0], "int"),
        ("PDIV 1次 不合格率", sa["rates"][0], sb["rates"][0], sb["rates"][0] - sa["rates"][0], "pct"),
        ("PDIV 2次 NOK 总数", sa["noks"][1], sb["noks"][1], sb["noks"][1] - sa["noks"][1], "int"),
        ("PDIV 2次 不合格率", sa["rates"][1], sb["rates"][1], sb["rates"][1] - sa["rates"][1], "pct"),
        ("PDIV 3次 NOK 总数", sa["noks"][2], sb["noks"][2], sb["noks"][2] - sa["noks"][2], "int"),
        ("PDIV 3次 不合格率", sa["rates"][2], sb["rates"][2], sb["rates"][2] - sa["rates"][2], "pct"),
        ("NOK 合计（1+2+3次）", sa["total_nok"], sb["total_nok"], sb["total_nok"] - sa["total_nok"], "int"),
        ("合计不合格率", sa["total_rate"], sb["total_rate"], sb["total_rate"] - sa["total_rate"], "pct"),
        ("单日最高不合格率", sa["max_rate"], sb["max_rate"], sb["max_rate"] - sa["max_rate"], "pct"),
        ("最高不合格率发生日", dates[sa["max_idx"]], dates[sb["max_idx"]], "—", "text"),
        ("单日最低不合格率", sa["min_rate"], sb["min_rate"], sb["min_rate"] - sa["min_rate"], "pct"),
        ("最低不合格率发生日", dates[sa["min_idx"]], dates[sb["min_idx"]], "—", "text"),
    ]

    for i, (label, va, vb, diff, kind) in enumerate(rows):
        r = 3 + i
        for col, value in enumerate((label, va, vb, diff), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border, cell.alignment = BORDER, CENTER
            if i == 0:
                cell.font, cell.fill = FONT_HEAD, FILL_SUM
            else:
                cell.font = FONT_BODY if col == 1 else FONT_TOTAL
                if i % 2 == 0:
                    cell.fill = FILL_ALT
                if col > 1:
                    cell.number_format = {"pct": "0.00%", "f1": "0.0", "int": "#,##0"}.get(kind, "General")
        ws.cell(row=r, column=1).alignment = LEFT
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 26
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    note_row = 3 + len(rows) + 1
    notes = [
        "说明：",
        "1. 不合格率 = 该次 PDIV 的 NOK 数量 ÷ 当日生产数量，由脚本按原始 NOK 数量重新计算，口径统一"
        "（原表 2026.07.27 A站 PDIV2 的 “2/0.95” 缺少百分号，此处按重算结果显示）。",
        "2. 合计不合格率 = (PDIV 1次 + 2次 + 3次 NOK) ÷ 当日生产数量，同一台可能被重复计入多次检测。",
        "3. 曲线图集中在「趋势图表」表；明细见「A站明细」「B站明细」，逐日对照见「A_B站对比」。",
        f"4. 本表由脚本 pdiv_nok_report_{VERSION}.py 生成。",
    ]
    for i, text in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1, value=text)
        cell.font = Font(name="微软雅黑", size=9, color="595959", bold=(i == 0))
        cell.alignment = LEFT


def write_raw_sheet(wb, src_path):
    """保留一份原始数据，便于回溯核对。"""
    src = load_workbook(src_path, data_only=True).worksheets[0]
    ws = wb.create_sheet("原始数据")
    for row in src.iter_rows(min_row=3, max_row=src.max_row, max_col=9, values_only=True):
        if all(v is None for v in row):
            continue
        ws.append(list(row))
    for col in range(1, 10):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.alignment = FONT_HEAD, CENTER
        cell.fill = PatternFill("solid", fgColor="808080")
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


# ---------------- 主流程 ----------------
def main(src_path, out_path):
    rows = read_raw(src_path)
    if not rows:
        print("原始表中没有读到数据行", file=sys.stderr)
        return 1

    dates = [r[0] for r in rows]
    data_a = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]
    data_b = [(r[0], r[5], r[6], r[7], r[8]) for r in rows]

    wb = Workbook()
    wb.remove(wb.active)
    ws_a = write_station_sheet(wb, SHEET_A, "A 站", data_a, FILL_A, RAMP_A)
    ws_b = write_station_sheet(wb, SHEET_B, "B 站", data_b, FILL_B, RAMP_B)
    write_compare_sheet(wb, data_a, data_b, dates)
    write_summary_sheet(wb, dates, data_a, data_b)
    write_raw_sheet(wb, src_path)
    chart_sheet = write_chart_sheet(wb, ws_a, ws_b, len(dates))
    wb.move_sheet(chart_sheet, offset=-(len(wb.sheetnames) - 1))  # 图表页放到第一位
    wb.active = 0
    wb.save(out_path)

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])
    n_charts = sum(len(ws._charts) for ws in wb)
    print(f"已生成：{out_path}（工作表 {len(wb.sheetnames)} 张，曲线图 {n_charts} 张）")
    print(f"生产日数 {len(dates)}（{dates[0]} ~ {dates[-1]}）")
    print(f"A 站：总产量 {sa['qty']}，NOK 合计 {sa['total_nok']}，合计不合格率 {sa['total_rate']:.2%}，"
          f"PDIV1/2/3 = {sa['rates'][0]:.2%} / {sa['rates'][1]:.2%} / {sa['rates'][2]:.2%}")
    print(f"B 站：总产量 {sb['qty']}，NOK 合计 {sb['total_nok']}，合计不合格率 {sb['total_rate']:.2%}，"
          f"PDIV1/2/3 = {sb['rates'][0]:.2%} / {sb['rates'][1]:.2%} / {sb['rates'][2]:.2%}")
    return 0


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("New_Microsoft_Excel_Worksheet.xlsx")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("PDIV检测数据统计_A站B站_v1.1.xlsx")
    sys.exit(main(src, out))
