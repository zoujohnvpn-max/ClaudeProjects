# -*- coding: utf-8 -*-
# =============================================
# Author: John
# Date: 2026-09-03
# Description: Split PDIV NOK raw data into station A/B sheets, compute statistics and build line charts
# =============================================
# --- Version History ---
# v1.0.0 (2026-09-03): initial release, split A/B stations, add summary sheet and line charts
# =============================================

VERSION = "v1.0.0"

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------- 样式常量 ----------------
FONT_TITLE = Font(name="微软雅黑", size=13, bold=True, color="1F3864")
FONT_HEAD = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="微软雅黑", size=10)
FONT_TOTAL = Font(name="微软雅黑", size=10, bold=True, color="1F3864")

FILL_A = PatternFill("solid", fgColor="2F5597")   # A 站表头（蓝）
FILL_B = PatternFill("solid", fgColor="C55A11")   # B 站表头（橙）
FILL_SUM = PatternFill("solid", fgColor="375623")  # 汇总表头（绿）
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")

# 曲线配色：A 站冷色系，B 站暖色系，同一 PDIV 次数在两站之间保持形状可辨
COLOR_A = ["2F5597", "4472C4", "8FAADC"]
COLOR_B = ["C55A11", "ED7D31", "F4B183"]

HEADERS = [
    "日期", "生产数量",
    "PDIV 1次 NOK", "PDIV 1次 不合格率",
    "PDIV 2次 NOK", "PDIV 2次 不合格率",
    "PDIV 3次 NOK", "PDIV 3次 不合格率",
    "NOK 合计", "合计不合格率",
]


def parse_nok(value):
    """把 '3/1.4%'、'0'、'2/0.95' 这类文本解析成 NOK 数量（整数）。"""
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0
    head = text.split("/")[0].strip()
    m = re.search(r"-?\d+(?:\.\d+)?", head)
    return int(float(m.group())) if m else 0


def parse_int(value):
    if value is None:
        return 0
    m = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return int(float(m.group())) if m else 0


def read_raw(path):
    """读取原始表，返回 [(日期, A站产量, A1, A2, A3, B站产量, B1, B2, B3), ...]。"""
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=4, max_col=9, values_only=True):
        date = row[0]
        if date is None or not str(date).strip():
            continue
        rows.append((
            str(date).strip(),
            parse_int(row[1]), parse_nok(row[2]), parse_nok(row[3]), parse_nok(row[4]),
            parse_int(row[5]), parse_nok(row[6]), parse_nok(row[7]), parse_nok(row[8]),
        ))
    return rows


def style_table(ws, head_fill, n_rows, n_cols=10):
    """统一表头 / 正文 / 合计行的样式和列宽。"""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, head_fill, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    for r in range(3, 3 + n_rows):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col in (4, 6, 8, 10):
                cell.number_format = "0.00%"

    total_row = 3 + n_rows
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_TOTAL, FILL_TOTAL, CENTER, BORDER
        if col in (4, 6, 8, 10):
            cell.number_format = "0.00%"

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 11
    for col in "CEGI":
        ws.column_dimensions[col].width = 15
    for col in "DFHJ":
        ws.column_dimensions[col].width = 17
    ws.freeze_panes = "B3"


def write_station_sheet(wb, title, station_name, data, head_fill, colors):
    """写入单个工位（A 站 / B 站）的明细表，并附三条 PDIV 不合格率曲线。"""
    ws = wb.create_sheet(title)
    ws["A1"] = f"{station_name} PDIV 检测数据明细（{data[0][0]} ~ {data[-1][0]}）"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.row_dimensions[1].height = 24

    for col, head in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col, value=head)

    for i, (date, qty, n1, n2, n3) in enumerate(data):
        r = 3 + i
        total_nok = n1 + n2 + n3
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=qty)
        for j, nok in enumerate((n1, n2, n3)):
            ws.cell(row=r, column=3 + j * 2, value=nok)
            ws.cell(row=r, column=4 + j * 2, value=(nok / qty if qty else 0))
        ws.cell(row=r, column=9, value=total_nok)
        ws.cell(row=r, column=10, value=(total_nok / qty if qty else 0))

    n = len(data)
    total_row = 3 + n
    ws.cell(row=total_row, column=1, value="合计 / 平均")
    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row - 1})")
    for col in (3, 5, 7, 9):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}3:{letter}{total_row - 1})")
        ws.cell(row=total_row, column=col + 1,
                value=f"=IF($B${total_row}=0,0,{letter}{total_row}/$B${total_row})")

    style_table(ws, head_fill, n)

    chart = LineChart()
    chart.title = f"{station_name} PDIV 不合格率趋势"
    chart.style = 2
    chart.height, chart.width = 9.5, 30
    chart.y_axis.title = "不合格率"
    chart.y_axis.numFmt = "0.0%"
    chart.x_axis.title = "日期"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    for j, col in enumerate((4, 6, 8)):
        ref = Reference(ws, min_col=col, min_row=2, max_row=total_row - 1)
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=total_row - 1))

    for j, series in enumerate(chart.series):
        series.smooth = False
        series.graphicalProperties.line = LineProperties(solidFill=colors[j], w=22000)
        series.marker = Marker(symbol="circle", size=5)
        series.marker.graphicalProperties.solidFill = colors[j]
        series.marker.graphicalProperties.line.solidFill = colors[j]

    ws.add_chart(chart, f"A{total_row + 3}")
    return ws


def write_compare_sheet(wb, data_a, data_b, dates):
    """A / B 站逐日对比明细 + 对比曲线（产量、各次 PDIV 不合格率）。"""
    ws = wb.create_sheet("A_B站对比")
    ws["A1"] = "A 站 / B 站 逐日对比"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 24

    heads = ["日期", "A站 产量", "B站 产量",
             "A站 PDIV1 不合格率", "B站 PDIV1 不合格率",
             "A站 PDIV2 不合格率", "B站 PDIV2 不合格率",
             "A站 合计不合格率", "B站 合计不合格率"]
    for col, head in enumerate(heads, start=1):
        cell = ws.cell(row=2, column=col, value=head)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, FILL_SUM, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    def rate(rec, idx):
        qty = rec[0]
        return (rec[idx] / qty) if qty else 0

    for i, date in enumerate(dates):
        r = 3 + i
        a, b = data_a[i][1:], data_b[i][1:]
        values = [date, a[0], b[0],
                  rate(a, 1), rate(b, 1),
                  rate(a, 2), rate(b, 2),
                  (sum(a[1:]) / a[0] if a[0] else 0), (sum(b[1:]) / b[0] if b[0] else 0)]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col >= 4:
                cell.number_format = "0.00%"

    last = 2 + len(dates)
    ws.column_dimensions["A"].width = 13
    for col in "BC":
        ws.column_dimensions[col].width = 12
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 20
    ws.freeze_panes = "B3"

    charts = [
        ("产量对比（A 站 vs B 站）", [2, 3], "台数", "General", ["2F5597", "C55A11"]),
        ("PDIV 1次 不合格率对比", [4, 5], "不合格率", "0.0%", ["2F5597", "C55A11"]),
        ("PDIV 2次 不合格率对比", [6, 7], "不合格率", "0.0%", ["4472C4", "ED7D31"]),
        ("合计不合格率对比（PDIV 1+2+3次）", [8, 9], "不合格率", "0.0%", ["1F3864", "843C0C"]),
    ]
    anchor = last + 3
    for title, cols, y_title, fmt, colors in charts:
        chart = LineChart()
        chart.title = title
        chart.style = 2
        chart.height, chart.width = 9.5, 30
        chart.y_axis.title = y_title
        chart.y_axis.numFmt = fmt
        chart.x_axis.title = "日期"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        for col in cols:
            chart.add_data(Reference(ws, min_col=col, min_row=2, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last))
        for j, series in enumerate(chart.series):
            series.smooth = False
            series.graphicalProperties.line = LineProperties(solidFill=colors[j], w=22000)
            series.marker = Marker(symbol="circle", size=5)
            series.marker.graphicalProperties.solidFill = colors[j]
            series.marker.graphicalProperties.line.solidFill = colors[j]
        ws.add_chart(chart, f"A{anchor}")
        anchor += 20
    return ws


def station_stats(records):
    """records: [(qty, n1, n2, n3), ...] -> 统计字典。"""
    qty = sum(r[0] for r in records)
    noks = [sum(r[i] for r in records) for i in (1, 2, 3)]
    total_nok = sum(noks)
    daily = [(sum(r[1:]) / r[0]) if r[0] else 0 for r in records]
    return {
        "days": len(records),
        "qty": qty,
        "noks": noks,
        "total_nok": total_nok,
        "rates": [n / qty if qty else 0 for n in noks],
        "total_rate": total_nok / qty if qty else 0,
        "max_rate": max(daily), "max_idx": daily.index(max(daily)),
        "min_rate": min(daily), "min_idx": daily.index(min(daily)),
        "avg_qty": qty / len(records) if records else 0,
    }


def write_summary_sheet(wb, dates, data_a, data_b):
    ws = wb.create_sheet("统计汇总", 0)
    ws["A1"] = f"PDIV 检测数据统计汇总（{dates[0]} ~ {dates[-1]}，共 {len(dates)} 个生产日）"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 26

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])

    rows = [
        ("统计项", "A 站", "B 站", "差值（B - A）", None),
        ("生产日数（天）", sa["days"], sb["days"], sb["days"] - sa["days"], "int"),
        ("总生产数量（台）", sa["qty"], sb["qty"], sb["qty"] - sa["qty"], "int"),
        ("日均生产数量（台）", sa["avg_qty"], sb["avg_qty"], sb["avg_qty"] - sa["avg_qty"], "f1"),
        ("PDIV 1次 NOK 总数", sa["noks"][0], sb["noks"][0], sb["noks"][0] - sa["noks"][0], "int"),
        ("PDIV 1次 不合格率", sa["rates"][0], sb["rates"][0], sb["rates"][0] - sa["rates"][0], "pct"),
        ("PDIV 2次 NOK 总数", sa["noks"][1], sb["noks"][1], sb["noks"][1] - sa["noks"][1], "int"),
        ("PDIV 2次 不合格率", sa["rates"][1], sb["rates"][1], sb["rates"][1] - sa["rates"][1], "pct"),
        ("PDIV 3次 NOK 总数", sa["noks"][2], sb["noks"][2], sb["noks"][2] - sa["noks"][2], "int"),
        ("PDIV 3次 不合格率", sa["rates"][2], sb["rates"][2], sb["rates"][2] - sa["rates"][2], "pct"),
        ("NOK 合计（1+2+3次）", sa["total_nok"], sb["total_nok"], sb["total_nok"] - sa["total_nok"], "int"),
        ("合计不合格率", sa["total_rate"], sb["total_rate"], sb["total_rate"] - sa["total_rate"], "pct"),
        ("单日最高不合格率", sa["max_rate"], sb["max_rate"], sb["max_rate"] - sa["max_rate"], "pct"),
        ("最高不合格率发生日", dates[sa["max_idx"]], dates[sb["max_idx"]], "—", "text"),
        ("单日最低不合格率", sa["min_rate"], sb["min_rate"], sb["min_rate"] - sa["min_rate"], "pct"),
        ("最低不合格率发生日", dates[sa["min_idx"]], dates[sb["min_idx"]], "—", "text"),
    ]

    for i, (label, va, vb, diff, kind) in enumerate(rows):
        r = 3 + i
        for col, value in enumerate((label, va, vb, diff), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border, cell.alignment = BORDER, CENTER
            if i == 0:
                cell.font, cell.fill = FONT_HEAD, FILL_SUM
            else:
                cell.font = FONT_BODY if col == 1 else FONT_TOTAL
                if i % 2 == 0:
                    cell.fill = FILL_ALT
                if col > 1:
                    if kind == "pct":
                        cell.number_format = "0.00%"
                    elif kind == "f1":
                        cell.number_format = "0.0"
                    elif kind == "int":
                        cell.number_format = "#,##0"
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 26
    for col in "BCD":
        ws.column_dimensions[col].width = 16

    note_row = 3 + len(rows) + 1
    notes = [
        "说明：",
        "1. 不合格率 = 该次 PDIV 的 NOK 数量 ÷ 当日生产数量，由脚本按原始 NOK 数量重新计算，"
        "口径统一（原表中 2026.07.27 A站 PDIV2 的 “2/0.95” 缺少百分号，此处按重算结果显示）。",
        "2. 合计不合格率 = (PDIV 1次 + 2次 + 3次 NOK) ÷ 当日生产数量，同一台可能被重复计入多次检测。",
        "3. A 站明细见「A站明细」表，B 站明细见「B站明细」表，逐日对比与曲线图见「A_B站对比」表。",
        f"4. 本表由脚本 pdiv_nok_report_{VERSION}.py 生成。",
    ]
    for i, text in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1, value=text)
        cell.font = Font(name="微软雅黑", size=9, color="595959", bold=(i == 0))
        cell.alignment = Alignment(horizontal="left", vertical="center")


def write_raw_sheet(wb, src_path):
    """保留一份原始数据，便于回溯核对。"""
    src = load_workbook(src_path, data_only=True).worksheets[0]
    ws = wb.create_sheet("原始数据")
    for row in src.iter_rows(min_row=3, max_row=src.max_row, max_col=9, values_only=True):
        if all(v is None for v in row):
            continue
        ws.append(list(row))
    for col in range(1, 10):
        ws.cell(row=1, column=col).font = FONT_HEAD
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="808080")
        ws.cell(row=1, column=col).alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def main(src_path, out_path):
    rows = read_raw(src_path)
    if not rows:
        print("原始表中没有读到数据行", file=sys.stderr)
        return 1

    dates = [r[0] for r in rows]
    data_a = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]
    data_b = [(r[0], r[5], r[6], r[7], r[8]) for r in rows]

    wb = Workbook()
    wb.remove(wb.active)
    write_station_sheet(wb, "A站明细", "A 站", data_a, FILL_A, COLOR_A)
    write_station_sheet(wb, "B站明细", "B 站", data_b, FILL_B, COLOR_B)
    write_compare_sheet(wb, data_a, data_b, dates)
    write_summary_sheet(wb, dates, data_a, data_b)
    write_raw_sheet(wb, src_path)
    wb.active = 0
    wb.save(out_path)

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])
    print(f"已生成：{out_path}")
    print(f"生产日数 {len(dates)}（{dates[0]} ~ {dates[-1]}）")
    print(f"A 站：总产量 {sa['qty']}，NOK 合计 {sa['total_nok']}，合计不合格率 {sa['total_rate']:.2%}，"
          f"PDIV1/2/3 = {sa['rates'][0]:.2%} / {sa['rates'][1]:.2%} / {sa['rates'][2]:.2%}")
    print(f"B 站：总产量 {sb['qty']}，NOK 合计 {sb['total_nok']}，合计不合格率 {sb['total_rate']:.2%}，"
          f"PDIV1/2/3 = {sb['rates'][0]:.2%} / {sb['rates'][1]:.2%} / {sb['rates'][2]:.2%}")
    return 0


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("New_Microsoft_Excel_Worksheet.xlsx")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(f"PDIV检测数据统计_A站B站_v1.0.xlsx")
    sys.exit(main(src, out))
