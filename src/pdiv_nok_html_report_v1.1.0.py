# -*- coding: utf-8 -*-
# =============================================
# Author: John
# Date: 2026-09-03
# Description: Build a standalone offline HTML trend report for PDIV NOK data of station A and B
# =============================================
# --- Version History ---
# v1.0.0 (2026-09-03): initial release, single-file HTML with interactive line charts and table view
# v1.1.0 (2026-09-03): round axis ticks to nice values, de-overlap end labels
# =============================================

VERSION = "v1.1.0"

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_nok(value):
    """把 '3/1.4%'、'0'、'2/0.95' 这类文本解析成 NOK 数量（整数）。"""
    if value is None:
        return 0
    head = str(value).strip().split("/")[0].strip()
    m = re.search(r"-?\d+(?:\.\d+)?", head)
    return int(float(m.group())) if m else 0


def parse_int(value):
    if value is None:
        return 0
    m = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return int(float(m.group())) if m else 0


def read_raw(path):
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=4, max_col=9, values_only=True):
        if row[0] is None or not str(row[0]).strip():
            continue
        rows.append((
            str(row[0]).strip(),
            parse_int(row[1]), parse_nok(row[2]), parse_nok(row[3]), parse_nok(row[4]),
            parse_int(row[5]), parse_nok(row[6]), parse_nok(row[7]), parse_nok(row[8]),
        ))
    return rows


HTML = r"""<!DOCTYPE html>
<html lang="zh-CN" data-palette="#2a78d6,#eb6834">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>PDIV 检测不合格率趋势 __VER__</title>
<style>
  :root{
    --surface-1:#fcfcfb; --plane:#f9f9f7;
    --text-primary:#0b0b0b; --text-secondary:#52514e; --text-muted:#898781;
    --grid:#e1e0d9; --axis:#c3c2b7; --border:#e6e5df;
    --a1:#184f95; --a2:#3987e5; --a3:#86b6ef;
    --b1:#a3400f; --b2:#eb6834; --b3:#f0906a;
    --sa:#2a78d6; --sb:#eb6834;
  }
  @media (prefers-color-scheme: dark){
    :root:not([data-theme="light"]){
      --surface-1:#1a1a19; --plane:#0d0d0d;
      --text-primary:#ffffff; --text-secondary:#c3c2b7; --text-muted:#898781;
      --grid:#2c2c2a; --axis:#383835; --border:#2c2c2a;
      --a1:#184f95; --a2:#3987e5; --a3:#86b6ef;
      --b1:#d95926; --b2:#eb6834; --b3:#f0906a;
      --sa:#3987e5; --sb:#d95926;
    }
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--plane);color:var(--text-primary);
       font:14px/1.55 "Segoe UI","Microsoft YaHei",system-ui,sans-serif}
  .wrap{max-width:1180px;margin:0 auto;padding:28px 20px 64px}
  h1{font-size:22px;margin:0 0 4px}
  .sub{color:var(--text-secondary);font-size:13px;margin:0 0 24px}
  .tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));gap:12px;margin-bottom:28px}
  .tile{background:var(--surface-1);border:1px solid var(--border);border-radius:10px;padding:14px 16px}
  .tile .k{font-size:12px;color:var(--text-secondary);margin-bottom:6px}
  .tile .v{font-size:26px;font-weight:650;letter-spacing:-.02em;line-height:1.1}
  .tile .n{font-size:12px;color:var(--text-muted);margin-top:4px}
  .card{background:var(--surface-1);border:1px solid var(--border);border-radius:12px;
        padding:18px 18px 10px;margin-bottom:20px;position:relative}
  .card h2{font-size:15px;margin:0 0 2px;font-weight:640}
  .card .cap{font-size:12px;color:var(--text-secondary);margin:0 0 10px}
  .legend{display:flex;flex-wrap:wrap;gap:16px;margin:0 0 6px;font-size:12px;color:var(--text-secondary)}
  .legend span{display:inline-flex;align-items:center;gap:6px}
  .swatch{width:14px;height:3px;border-radius:2px;display:inline-block}
  svg{display:block;width:100%;height:auto;overflow:visible}
  .tip{position:absolute;pointer-events:none;opacity:0;transition:opacity .08s;
       background:var(--surface-1);border:1px solid var(--axis);border-radius:8px;
       padding:8px 10px;font-size:12px;box-shadow:0 4px 14px rgba(0,0,0,.14);z-index:5;white-space:nowrap}
  .tip b{display:block;margin-bottom:4px;color:var(--text-primary);font-weight:640}
  .tip i{font-style:normal;color:var(--text-secondary)}
  .tip .row{display:flex;align-items:center;gap:6px}
  .tip .dot{width:8px;height:8px;border-radius:50%;flex:0 0 auto}
  details{margin-top:24px}
  summary{cursor:pointer;font-size:13px;color:var(--text-secondary);padding:6px 0}
  table{border-collapse:collapse;width:100%;font-size:12px;margin-top:8px}
  th,td{border:1px solid var(--border);padding:5px 8px;text-align:center;white-space:nowrap}
  th{background:var(--plane);color:var(--text-secondary);font-weight:600}
  .scroll{overflow-x:auto}
  footer{color:var(--text-muted);font-size:12px;margin-top:28px}
</style>
</head>
<body>
<div class="wrap">
  <h1>PDIV 检测不合格率趋势</h1>
  <p class="sub" id="sub"></p>
  <div class="tiles" id="tiles"></div>
  <div id="charts"></div>

  <details>
    <summary>展开明细数据表（A 站 / B 站 逐日）</summary>
    <div class="scroll"><table id="tbl"></table></div>
  </details>

  <footer id="foot"></footer>
</div>

<script>
const DATA = __DATA__;
const pct = v => (v*100).toFixed(2) + '%';
const cssv = n => getComputedStyle(document.documentElement).getPropertyValue(n).trim();

function rate(st, k, i){ return st.qty[i] ? st['n'+k][i] / st.qty[i] : 0; }
function totalRate(st, i){ return st.qty[i] ? (st.n1[i]+st.n2[i]+st.n3[i]) / st.qty[i] : 0; }

/* ---------- 折线图 ---------- */
function chart(host, cfg){
  const W=1080, H=330, M={t:16,r:88,b:52,l:64};
  const pw=W-M.l-M.r, ph=H-M.t-M.b;
  const cats=DATA.dates, n=cats.length;
  const all=cfg.series.flatMap(s=>s.values);
  const min=Math.min(0,...all);
  // 轴上界取整到 1/2/2.5/5 的整数倍，刻度读数才是整齐的 2% / 4% / 6%
  const nice=v=>{ if(v<=0) return 1;
    const e=Math.pow(10,Math.floor(Math.log10(v))), f=v/e;
    return (f<=1?1:f<=2?2:f<=2.5?2.5:f<=5?5:10)*e; };
  const max=nice(Math.max(...all)*1.05);
  const x=i=> M.l + (n===1?pw/2:pw*i/(n-1));
  const y=v=> M.t + ph - ph*(v-min)/(max-min);
  const ticks=5, fmt=cfg.fmt;

  const NS='http://www.w3.org/2000/svg';
  const el=(t,a)=>{const e=document.createElementNS(NS,t);for(const k in a)e.setAttribute(k,a[k]);return e;};
  const svg=el('svg',{viewBox:`0 0 ${W} ${H}`,role:'img','aria-label':cfg.title});

  for(let i=0;i<=ticks;i++){
    const v=min+(max-min)*i/ticks, yy=y(v);
    svg.appendChild(el('line',{x1:M.l,x2:M.l+pw,y1:yy,y2:yy,stroke:cssv('--grid'),'stroke-width':1}));
    const t=el('text',{x:M.l-10,y:yy+4,'text-anchor':'end','font-size':11,fill:cssv('--text-muted')});
    t.textContent=fmt(v); svg.appendChild(t);
  }
  svg.appendChild(el('line',{x1:M.l,x2:M.l+pw,y1:M.t+ph,y2:M.t+ph,stroke:cssv('--axis'),'stroke-width':1}));

  const step=Math.ceil(n/12);
  cats.forEach((c,i)=>{ if(i%step) return;
    const t=el('text',{x:x(i),y:M.t+ph+20,'text-anchor':'end','font-size':10.5,
                       fill:cssv('--text-muted'),transform:`rotate(-40 ${x(i)} ${M.t+ph+20})`});
    t.textContent=c.slice(5); svg.appendChild(t);
  });
  const xl=el('text',{x:M.l+pw/2,y:H-4,'text-anchor':'middle','font-size':11,fill:cssv('--text-muted')});
  xl.textContent='日期（月.日）'; svg.appendChild(xl);

  const crosshair=el('line',{y1:M.t,y2:M.t+ph,stroke:cssv('--axis'),'stroke-width':1,opacity:0});
  svg.appendChild(crosshair);

  const endLabels=[];
  cfg.series.forEach(s=>{
    const color=cssv(s.color);
    const d=s.values.map((v,i)=>`${i?'L':'M'}${x(i).toFixed(1)} ${y(v).toFixed(1)}`).join(' ');
    svg.appendChild(el('path',{d,fill:'none',stroke:color,'stroke-width':2,
                               'stroke-linejoin':'round','stroke-linecap':'round'}));
    s.values.forEach((v,i)=>{
      svg.appendChild(el('circle',{cx:x(i),cy:y(v),r:3.4,fill:color,
                                   stroke:cssv('--surface-1'),'stroke-width':1.6}));
    });
    // 末端直接标注，识别不只靠颜色；位置先记下，稍后统一避让
    const lbl=el('text',{x:M.l+pw+8,'font-size':11.5,fill:cssv('--text-secondary')});
    lbl.textContent=s.name; svg.appendChild(lbl);
    endLabels.push({node:lbl, y:y(s.values[n-1])+4});
  });

  // 末端标签互相压住时按 14px 间距上下推开
  endLabels.sort((p,q)=>p.y-q.y);
  for(let i=1;i<endLabels.length;i++){
    const gap=endLabels[i].y-endLabels[i-1].y;
    if(gap<14) endLabels[i].y=endLabels[i-1].y+14;
  }
  endLabels.forEach(l=>l.node.setAttribute('y',l.y));

  const hit=el('rect',{x:M.l,y:M.t,width:pw,height:ph,fill:'transparent'});
  svg.appendChild(hit);

  const tip=document.createElement('div'); tip.className='tip'; host.appendChild(tip);
  hit.addEventListener('mousemove',ev=>{
    const box=svg.getBoundingClientRect();
    const px=(ev.clientX-box.left)/box.width*W;
    const i=Math.max(0,Math.min(n-1,Math.round((px-M.l)/(pw/(n-1)))));
    crosshair.setAttribute('x1',x(i)); crosshair.setAttribute('x2',x(i));
    crosshair.setAttribute('opacity',1);
    tip.innerHTML='<b>'+cats[i]+'</b>'+cfg.series.map(s=>
      `<div class="row"><span class="dot" style="background:${cssv(s.color)}"></span>`+
      `<i>${s.name}</i>&nbsp;<span>${fmt(s.values[i])}</span></div>`).join('');
    tip.style.opacity=1;
    const hb=host.getBoundingClientRect();
    let left=ev.clientX-hb.left+14;
    if(left+tip.offsetWidth>hb.width-8) left=ev.clientX-hb.left-tip.offsetWidth-14;
    tip.style.left=left+'px';
    tip.style.top=(ev.clientY-hb.top-tip.offsetHeight-10)+'px';
  });
  hit.addEventListener('mouseleave',()=>{tip.style.opacity=0;crosshair.setAttribute('opacity',0);});

  host.appendChild(svg);
}

function card(title, cap, series, fmt){
  const c=document.createElement('div'); c.className='card';
  c.innerHTML=`<h2>${title}</h2><p class="cap">${cap}</p>`;
  if(series.length>1){
    c.innerHTML+='<div class="legend">'+series.map(s=>
      `<span><i class="swatch" style="background:var(${s.color})"></i>${s.name}</span>`).join('')+'</div>';
  }
  document.getElementById('charts').appendChild(c);
  chart(c,{title,series,fmt});
}

/* ---------- 渲染 ---------- */
const A=DATA.A, B=DATA.B, N=DATA.dates.length;
const idx=[...Array(N).keys()];
const sum=a=>a.reduce((x,y)=>x+y,0);
const stat=st=>{
  const q=sum(st.qty), nk=[sum(st.n1),sum(st.n2),sum(st.n3)];
  return {qty:q, noks:nk, tot:nk[0]+nk[1]+nk[2], rate:(nk[0]+nk[1]+nk[2])/q,
          r1:nk[0]/q, r2:nk[1]/q, r3:nk[2]/q};
};
const SA=stat(A), SB=stat(B);

document.getElementById('sub').textContent =
  `${DATA.dates[0]} ~ ${DATA.dates[N-1]}，共 ${N} 个生产日 · 不合格率 = 当次 PDIV NOK 数 ÷ 当日生产数量`;

document.getElementById('tiles').innerHTML = [
  ['A 站 合计不合格率', pct(SA.rate), `产量 ${SA.qty} 台 · NOK ${SA.tot}`],
  ['B 站 合计不合格率', pct(SB.rate), `产量 ${SB.qty} 台 · NOK ${SB.tot}`],
  ['A 站 PDIV 1次 不合格率', pct(SA.r1), `NOK ${SA.noks[0]}`],
  ['B 站 PDIV 1次 不合格率', pct(SB.r1), `NOK ${SB.noks[0]}`],
  ['B 站 − A 站 差值', '+'+((SB.rate-SA.rate)*100).toFixed(2)+' pp', 'B 站约为 A 站的 '+(SB.rate/SA.rate).toFixed(1)+' 倍'],
].map(([k,v,n])=>`<div class="tile"><div class="k">${k}</div><div class="v">${v}</div><div class="n">${n}</div></div>`).join('');

card('A 站 · PDIV 1次 不合格率','单条曲线，看 A 站首检不合格率的日间波动。',
     [{name:'A站 PDIV 1次',color:'--sa',values:idx.map(i=>rate(A,1,i))}],pct);
card('B 站 · PDIV 1次 不合格率','同口径的 B 站首检不合格率。',
     [{name:'B站 PDIV 1次',color:'--sb',values:idx.map(i=>rate(B,1,i))}],pct);
card('A 站 · PDIV 1次 / 2次 / 3次 不合格率','同一工位内三次检测的不合格率，颜色由深到浅对应第 1 / 2 / 3 次。',
     [{name:'PDIV 1次',color:'--a1',values:idx.map(i=>rate(A,1,i))},
      {name:'PDIV 2次',color:'--a2',values:idx.map(i=>rate(A,2,i))},
      {name:'PDIV 3次',color:'--a3',values:idx.map(i=>rate(A,3,i))}],pct);
card('B 站 · PDIV 1次 / 2次 / 3次 不合格率','同上，B 站。',
     [{name:'PDIV 1次',color:'--b1',values:idx.map(i=>rate(B,1,i))},
      {name:'PDIV 2次',color:'--b2',values:idx.map(i=>rate(B,2,i))},
      {name:'PDIV 3次',color:'--b3',values:idx.map(i=>rate(B,3,i))}],pct);
card('A 站 vs B 站 · PDIV 1次 不合格率','两站首检直接对比，差距最主要来自这一项。',
     [{name:'A 站',color:'--sa',values:idx.map(i=>rate(A,1,i))},
      {name:'B 站',color:'--sb',values:idx.map(i=>rate(B,1,i))}],pct);
card('A 站 vs B 站 · 合计不合格率','PDIV 1+2+3 次合计。',
     [{name:'A 站',color:'--sa',values:idx.map(i=>totalRate(A,i))},
      {name:'B 站',color:'--sb',values:idx.map(i=>totalRate(B,i))}],pct);
card('A 站 vs B 站 · 每日生产数量','两站产量基本同步，说明不合格率差异不是产量差异带来的。',
     [{name:'A 站',color:'--sa',values:A.qty},
      {name:'B 站',color:'--sb',values:B.qty}],v=>Math.round(v));

const head=['日期','A站 产量','A站 P1','A站 P1率','A站 P2','A站 P2率','A站 P3','A站 P3率','A站 合计率',
            'B站 产量','B站 P1','B站 P1率','B站 P2','B站 P2率','B站 P3','B站 P3率','B站 合计率'];
document.getElementById('tbl').innerHTML =
  '<thead><tr>'+head.map(h=>`<th>${h}</th>`).join('')+'</tr></thead><tbody>'+
  idx.map(i=>'<tr>'+[DATA.dates[i],
    A.qty[i],A.n1[i],pct(rate(A,1,i)),A.n2[i],pct(rate(A,2,i)),A.n3[i],pct(rate(A,3,i)),pct(totalRate(A,i)),
    B.qty[i],B.n1[i],pct(rate(B,1,i)),B.n2[i],pct(rate(B,2,i)),B.n3[i],pct(rate(B,3,i)),pct(totalRate(B,i))
  ].map(v=>`<td>${v}</td>`).join('')+'</tr>').join('')+'</tbody>';

document.getElementById('foot').textContent =
  '本页由脚本 pdiv_nok_html_report___VER__.py 生成，离线单文件，双击即可用浏览器打开。';
</script>
</body>
</html>
"""


def main(src_path, out_path):
    rows = read_raw(src_path)
    if not rows:
        print("原始表中没有读到数据行", file=sys.stderr)
        return 1

    payload = {
        "dates": [r[0] for r in rows],
        "A": {"qty": [r[1] for r in rows], "n1": [r[2] for r in rows],
              "n2": [r[3] for r in rows], "n3": [r[4] for r in rows]},
        "B": {"qty": [r[5] for r in rows], "n1": [r[6] for r in rows],
              "n2": [r[7] for r in rows], "n3": [r[8] for r in rows]},
    }
    html = HTML.replace("__DATA__", json.dumps(payload, ensure_ascii=False)).replace("__VER__", VERSION)
    Path(out_path).write_text(html, encoding="utf-8")
    print(f"已生成：{out_path}（{len(rows)} 个生产日，7 张曲线图）")
    return 0


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("New_Microsoft_Excel_Worksheet.xlsx")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("PDIV检测数据趋势_v1.0.html")
    sys.exit(main(src, out))
