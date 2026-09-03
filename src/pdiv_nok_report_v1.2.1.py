# -*- coding: utf-8 -*-
# =============================================
# Author: John
# Date: 2026-09-03
# Description: Split PDIV NOK raw data into station A/B sheets, compute statistics and build line charts
# =============================================
# --- Version History ---
# v1.0.0 (2026-09-03): initial release, split A/B stations, add summary sheet and line charts
# v1.1.0 (2026-09-03): add dedicated chart sheet with per-metric line charts, fix category axis
#                      position, use string references for date categories, apply validated palette
# v1.2.0 (2026-09-03): add --lang switch, ship English wording for every sheet, header and note
# v1.2.1 (2026-09-03): reword English raw-data headers and note 1 after a translation review
# =============================================

VERSION = "v1.2.1"

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------- 文案表（中 / 英） ----------------
TEXT = {
    "zh": {
        "font": "微软雅黑",
        "sheet_chart": "趋势图表", "sheet_a": "A站明细", "sheet_b": "B站明细",
        "sheet_cmp": "A_B站对比", "sheet_sum": "统计汇总", "sheet_raw": "原始数据",
        "station_a": "A 站", "station_b": "B 站",
        "headers": ["日期", "生产数量",
                    "PDIV 1次 NOK", "PDIV 1次 不合格率", "PDIV 2次 NOK", "PDIV 2次 不合格率",
                    "PDIV 3次 NOK", "PDIV 3次 不合格率", "NOK 合计", "合计不合格率"],
        "detail_title": "{st} PDIV 检测数据明细（{d0} ~ {d1}）",
        "total_row": "合计 / 平均",
        "trend_title": "{st} PDIV 不合格率趋势",
        "axis_date": "日期", "axis_rate": "不合格率", "axis_qty": "台数",
        "chart_page_title": "PDIV 检测不合格率趋势图",
        "chart_page_note": "数据源：{a} / {b} 两张表，改动明细数据后图表自动更新。",
        "c_p1": "{st} · PDIV 1次 不合格率",
        "c_p123": "{st} · PDIV 1次 / 2次 / 3次 不合格率",
        "c_total": "{st} · 合计不合格率",
        "c_qty": "{st} · 每日生产数量",
        "cmp_title": "A 站 / B 站 逐日对比",
        "cmp_heads": ["日期", "A站 产量", "B站 产量",
                      "A站 PDIV1 不合格率", "B站 PDIV1 不合格率",
                      "A站 PDIV2 不合格率", "B站 PDIV2 不合格率",
                      "A站 合计不合格率", "B站 合计不合格率"],
        "cmp_c1": "PDIV 1次 不合格率对比（A 站 vs B 站）",
        "cmp_c2": "PDIV 2次 不合格率对比（A 站 vs B 站）",
        "cmp_c3": "合计不合格率对比（PDIV 1+2+3次）",
        "cmp_c4": "每日产量对比（A 站 vs B 站）",
        "sum_title": "PDIV 检测数据统计汇总（{d0} ~ {d1}，共 {n} 个生产日）",
        "sum_heads": ["统计项", "A 站", "B 站", "差值（B - A）"],
        "sum_rows": ["生产日数（天）", "总生产数量（台）", "日均生产数量（台）",
                     "PDIV 1次 NOK 总数", "PDIV 1次 不合格率",
                     "PDIV 2次 NOK 总数", "PDIV 2次 不合格率",
                     "PDIV 3次 NOK 总数", "PDIV 3次 不合格率",
                     "NOK 合计（1+2+3次）", "合计不合格率",
                     "单日最高不合格率", "最高不合格率发生日",
                     "单日最低不合格率", "最低不合格率发生日"],
        "notes": [
            "说明：",
            "1. 不合格率 = 该次 PDIV 的 NOK 数量 ÷ 当日生产数量，由脚本按原始 NOK 数量重新计算，口径统一"
            "（原表 2026.07.27 A站 PDIV2 的 “2/0.95” 缺少百分号，此处按重算结果显示）。",
            "2. 合计不合格率 = (PDIV 1次 + 2次 + 3次 NOK) ÷ 当日生产数量，同一台可能被重复计入多次检测。",
            "3. 曲线图集中在「{chart}」表；明细见「{a}」「{b}」，逐日对照见「{cmp}」。",
            "4. 本表由脚本 pdiv_nok_report_{ver}.py 生成。",
        ],
        "marker": "▼ ",
        "raw_heads": None,
    },
    "en": {
        "font": "Segoe UI",
        "sheet_chart": "Charts", "sheet_a": "Station A", "sheet_b": "Station B",
        "sheet_cmp": "A vs B", "sheet_sum": "Summary", "sheet_raw": "Raw Data",
        "station_a": "Station A", "station_b": "Station B",
        "headers": ["Date", "Output Qty",
                    "PDIV #1 NOK", "PDIV #1 Reject Rate", "PDIV #2 NOK", "PDIV #2 Reject Rate",
                    "PDIV #3 NOK", "PDIV #3 Reject Rate", "Total NOK", "Total Reject Rate"],
        "detail_title": "{st} — PDIV Inspection Data ({d0} – {d1})",
        "total_row": "Total / Average",
        "trend_title": "{st} — PDIV Reject Rate Trend",
        "axis_date": "Date", "axis_rate": "Reject rate", "axis_qty": "Units",
        "chart_page_title": "PDIV Reject Rate Trend Charts",
        "chart_page_note": "Source: sheets \"{a}\" and \"{b}\". Charts update automatically when the data changes.",
        "c_p1": "{st} — PDIV #1 Reject Rate",
        "c_p123": "{st} — PDIV #1 / #2 / #3 Reject Rate",
        "c_total": "{st} — Total Reject Rate",
        "c_qty": "{st} — Daily Output",
        "cmp_title": "Station A vs Station B — Daily Comparison",
        "cmp_heads": ["Date", "A Output", "B Output",
                      "A PDIV #1 Reject Rate", "B PDIV #1 Reject Rate",
                      "A PDIV #2 Reject Rate", "B PDIV #2 Reject Rate",
                      "A Total Reject Rate", "B Total Reject Rate"],
        "cmp_c1": "PDIV #1 Reject Rate — Station A vs Station B",
        "cmp_c2": "PDIV #2 Reject Rate — Station A vs Station B",
        "cmp_c3": "Total Reject Rate (PDIV #1+#2+#3) — Station A vs Station B",
        "cmp_c4": "Daily Output — Station A vs Station B",
        "sum_title": "PDIV Inspection Statistics Summary ({d0} – {d1}, {n} production days)",
        "sum_heads": ["Metric", "Station A", "Station B", "Difference (B − A)"],
        "sum_rows": ["Production days", "Total output (units)", "Average daily output (units)",
                     "PDIV #1 NOK total", "PDIV #1 reject rate",
                     "PDIV #2 NOK total", "PDIV #2 reject rate",
                     "PDIV #3 NOK total", "PDIV #3 reject rate",
                     "Total NOK (#1+#2+#3)", "Total reject rate",
                     "Highest daily reject rate", "Date of highest reject rate",
                     "Lowest daily reject rate", "Date of lowest reject rate"],
        "notes": [
            "Notes:",
            "1. Reject rate = NOK count of that PDIV pass / output of the same day. Rates are recalculated "
            "by the script from the raw NOK counts so the basis is consistent, and the values shown "
            "here are those recalculated figures "
            "(the raw file wrote \"2/0.95\" without a percent sign for Station A PDIV #2 on 2026.07.27).",
            "2. Total reject rate = (PDIV #1 + #2 + #3 NOK) / daily output; one unit may be counted in "
            "more than one inspection pass.",
            "3. All line charts are on the \"{chart}\" sheet; detail data on \"{a}\" and \"{b}\"; "
            "day-by-day comparison on \"{cmp}\".",
            "4. Generated by pdiv_nok_report_{ver}.py.",
        ],
        "marker": "▼ ",
        # 原始表的表头是中文，英文版把这一行译过来，数据本身原样保留
        "raw_heads": ["Date", "Station A total output",
                      "Station A PDIV #1 NOK / rate", "Station A PDIV #2 NOK / rate",
                      "Station A PDIV #3 NOK / rate", "Station B total output",
                      "Station B PDIV #1 NOK / rate", "Station B PDIV #2 NOK / rate",
                      "Station B PDIV #3 NOK / rate"],
    },
}

T = TEXT["zh"]  # 由 set_lang() 切换

# ---------------- 配色（经 dataviz validate_palette 校验） ----------------
# A 站 = 蓝色序数梯度，B 站 = 橙色序数梯度，PDIV 次数越靠前颜色越深；
# A/B 对比图用蓝 / 橙两个分类色，CVD 与常规视觉分离度均达标。
RAMP_A = ["184F95", "3987E5", "86B6EF"]
RAMP_B = ["A3400F", "EB6834", "F0906A"]
COLOR_A, COLOR_B = "2A78D6", "EB6834"

FILL_A = PatternFill("solid", fgColor="184F95")
FILL_B = PatternFill("solid", fgColor="A3400F")
FILL_SUM = PatternFill("solid", fgColor="375623")
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

FONT_TITLE = FONT_SUB = FONT_HEAD = FONT_BODY = FONT_TOTAL = FONT_NOTE = None


def set_lang(lang):
    """切换语言，同时按语言选择字体（中文用微软雅黑，英文用 Segoe UI）。"""
    global T, FONT_TITLE, FONT_SUB, FONT_HEAD, FONT_BODY, FONT_TOTAL, FONT_NOTE
    T = TEXT[lang]
    f = T["font"]
    FONT_TITLE = Font(name=f, size=13, bold=True, color="1F3864")
    FONT_SUB = Font(name=f, size=11, bold=True, color="52514E")
    FONT_HEAD = Font(name=f, size=10, bold=True, color="FFFFFF")
    FONT_BODY = Font(name=f, size=10)
    FONT_TOTAL = Font(name=f, size=10, bold=True, color="1F3864")
    FONT_NOTE = Font(name=f, size=9, color="595959")


# ---------------- 原始数据解析 ----------------
def parse_nok(value):
    """把 '3/1.4%'、'0'、'2/0.95' 这类文本解析成 NOK 数量（整数）。"""
    if value is None:
        return 0
    head = str(value).strip().split("/")[0].strip()
    m = re.search(r"-?\d+(?:\.\d+)?", head)
    return int(float(m.group())) if m else 0


def parse_int(value):
    if value is None:
        return 0
    m = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return int(float(m.group())) if m else 0


def read_raw(path):
    """读取原始表，返回 [(日期, A站产量, A1, A2, A3, B站产量, B1, B2, B3), ...]。"""
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=4, max_col=9, values_only=True):
        if row[0] is None or not str(row[0]).strip():
            continue
        rows.append((
            str(row[0]).strip(),
            parse_int(row[1]), parse_nok(row[2]), parse_nok(row[3]), parse_nok(row[4]),
            parse_int(row[5]), parse_nok(row[6]), parse_nok(row[7]), parse_nok(row[8]),
        ))
    return rows


# ---------------- 图表工厂 ----------------
def make_line_chart(title, y_title, num_fmt, series_defs, cat_sheet, cat_range, width=26, height=9):
    """按统一规范生成折线图。

    series_defs: [(数据源 worksheet, 列号, 线条颜色), ...]，系列名取该列第 2 行表头。
    cat_range:   分类轴（日期）区间，形如 "$A$3:$A$33"。
    """
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height, chart.width = height, width
    chart.y_axis.title = y_title
    chart.y_axis.numFmt = num_fmt
    chart.x_axis.title = T["axis_date"]
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.axPos = "b"          # 分类轴必须在下方，openpyxl 默认值是 "l"
    chart.y_axis.axPos = "l"
    chart.x_axis.tickLblSkip = 2      # 31 个日期，隔一个显示，避免标签重叠
    chart.x_axis.tickMarkSkip = 2

    last_row = int(re.search(r"\d+$", cat_range).group())
    for ws, col, _ in series_defs:
        chart.add_data(Reference(ws, min_col=col, min_row=2, max_row=last_row), titles_from_data=True)

    # 日期是文本，用 StrRef 而不是 openpyxl 默认的 NumRef，否则部分阅读器把日期当数字
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f="'{}'!{}".format(cat_sheet, cat_range)))

    for series, (_, _, color) in zip(chart.series, series_defs):
        series.smooth = False
        series.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
        series.marker = Marker(symbol="circle", size=6)
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color

    # 单条曲线不需要图例，标题已说明是哪条；两条以上必须有图例
    if len(series_defs) == 1:
        chart.legend = None
    else:
        chart.legend.position = "b"
        chart.legend.overlay = False
    return chart


# ---------------- 明细表 ----------------
def style_table(ws, head_fill, n_rows, n_cols=10):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, head_fill, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    for r in range(3, 3 + n_rows):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col in (4, 6, 8, 10):
                cell.number_format = "0.00%"

    total_row = 3 + n_rows
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_TOTAL, FILL_TOTAL, CENTER, BORDER
        if col in (4, 6, 8, 10):
            cell.number_format = "0.00%"

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 12
    for col in "CEGI":
        ws.column_dimensions[col].width = 16
    for col in "DFHJ":
        ws.column_dimensions[col].width = 19
    ws.freeze_panes = "B3"


def write_station_sheet(wb, sheet_title, station_name, data, head_fill, ramp):
    """单个工位的明细表，表格下方附该工位的三条 PDIV 不合格率曲线。"""
    ws = wb.create_sheet(sheet_title)
    ws["A1"] = T["detail_title"].format(st=station_name, d0=data[0][0], d1=data[-1][0])
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.row_dimensions[1].height = 24

    for col, head in enumerate(T["headers"], start=1):
        ws.cell(row=2, column=col, value=head)

    for i, (date, qty, n1, n2, n3) in enumerate(data):
        r = 3 + i
        total_nok = n1 + n2 + n3
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=qty)
        for j, nok in enumerate((n1, n2, n3)):
            ws.cell(row=r, column=3 + j * 2, value=nok)
            ws.cell(row=r, column=4 + j * 2, value=(nok / qty if qty else 0))
        ws.cell(row=r, column=9, value=total_nok)
        ws.cell(row=r, column=10, value=(total_nok / qty if qty else 0))

    n = len(data)
    total_row = 3 + n
    ws.cell(row=total_row, column=1, value=T["total_row"])
    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{total_row - 1})")
    for col in (3, 5, 7, 9):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}3:{letter}{total_row - 1})")
        ws.cell(row=total_row, column=col + 1,
                value=f"=IF($B${total_row}=0,0,{letter}{total_row}/$B${total_row})")

    style_table(ws, head_fill, n)

    title = T["trend_title"].format(st=station_name)
    ws.cell(row=total_row + 2, column=1, value=T["marker"] + title).font = FONT_SUB
    ws.add_chart(make_line_chart(title, T["axis_rate"], "0.0%",
                                 [(ws, 4, ramp[0]), (ws, 6, ramp[1]), (ws, 8, ramp[2])],
                                 sheet_title, f"$A$3:$A${total_row - 1}"),
                 f"A{total_row + 3}")
    return ws


# ---------------- 图表专页 ----------------
def write_chart_sheet(wb, ws_a, ws_b, n_days):
    """所有曲线图集中放在这张表，打开工作簿第一眼就能看到。"""
    ws = wb.create_sheet(T["sheet_chart"])
    cat = f"$A$3:$A${2 + n_days}"
    sa, sb = T["station_a"], T["station_b"]

    ws["A1"] = T["chart_page_title"]
    ws["A1"].font = FONT_TITLE
    ws["A2"] = T["chart_page_note"].format(a=T["sheet_a"], b=T["sheet_b"])
    ws["A2"].font = FONT_NOTE
    ws.column_dimensions["A"].width = 4

    specs = [
        (T["c_p1"].format(st=sa), [(ws_a, 4, COLOR_A)], T["sheet_a"], T["axis_rate"], "0.0%"),
        (T["c_p1"].format(st=sb), [(ws_b, 4, COLOR_B)], T["sheet_b"], T["axis_rate"], "0.0%"),
        (T["c_p123"].format(st=sa),
         [(ws_a, 4, RAMP_A[0]), (ws_a, 6, RAMP_A[1]), (ws_a, 8, RAMP_A[2])],
         T["sheet_a"], T["axis_rate"], "0.0%"),
        (T["c_p123"].format(st=sb),
         [(ws_b, 4, RAMP_B[0]), (ws_b, 6, RAMP_B[1]), (ws_b, 8, RAMP_B[2])],
         T["sheet_b"], T["axis_rate"], "0.0%"),
        (T["c_total"].format(st=sa), [(ws_a, 10, COLOR_A)], T["sheet_a"], T["axis_rate"], "0.0%"),
        (T["c_total"].format(st=sb), [(ws_b, 10, COLOR_B)], T["sheet_b"], T["axis_rate"], "0.0%"),
        (T["c_qty"].format(st=sa), [(ws_a, 2, COLOR_A)], T["sheet_a"], T["axis_qty"], "General"),
        (T["c_qty"].format(st=sb), [(ws_b, 2, COLOR_B)], T["sheet_b"], T["axis_qty"], "General"),
    ]

    row = 4
    for title, series_defs, cat_sheet, y_title, fmt in specs:
        ws.cell(row=row, column=1, value=T["marker"] + title).font = FONT_SUB
        ws.add_chart(make_line_chart(title, y_title, fmt, series_defs, cat_sheet, cat), f"A{row + 1}")
        row += 21
    return ws


def write_compare_sheet(wb, data_a, data_b, dates):
    """A / B 站逐日对比明细 + 对比曲线。"""
    ws = wb.create_sheet(T["sheet_cmp"])
    ws["A1"] = T["cmp_title"]
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 24

    for col, head in enumerate(T["cmp_heads"], start=1):
        cell = ws.cell(row=2, column=col, value=head)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEAD, FILL_SUM, CENTER, BORDER
    ws.row_dimensions[2].height = 32

    def rate(rec, idx):
        return (rec[idx] / rec[0]) if rec[0] else 0

    for i, date in enumerate(dates):
        r = 3 + i
        a, b = data_a[i][1:], data_b[i][1:]
        values = [date, a[0], b[0], rate(a, 1), rate(b, 1), rate(a, 2), rate(b, 2),
                  (sum(a[1:]) / a[0] if a[0] else 0), (sum(b[1:]) / b[0] if b[0] else 0)]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font, cell.alignment, cell.border = FONT_BODY, CENTER, BORDER
            if (r - 3) % 2 == 1:
                cell.fill = FILL_ALT
            if col >= 4:
                cell.number_format = "0.00%"

    last = 2 + len(dates)
    ws.column_dimensions["A"].width = 13
    for col in "BC":
        ws.column_dimensions[col].width = 13
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 22
    ws.freeze_panes = "B3"

    cat = f"$A$3:$A${last}"
    charts = [
        (T["cmp_c1"], [(ws, 4, COLOR_A), (ws, 5, COLOR_B)], T["axis_rate"], "0.0%"),
        (T["cmp_c2"], [(ws, 6, COLOR_A), (ws, 7, COLOR_B)], T["axis_rate"], "0.0%"),
        (T["cmp_c3"], [(ws, 8, COLOR_A), (ws, 9, COLOR_B)], T["axis_rate"], "0.0%"),
        (T["cmp_c4"], [(ws, 2, COLOR_A), (ws, 3, COLOR_B)], T["axis_qty"], "General"),
    ]
    row = last + 2
    for title, series_defs, y_title, fmt in charts:
        ws.cell(row=row, column=1, value=T["marker"] + title).font = FONT_SUB
        ws.add_chart(make_line_chart(title, y_title, fmt, series_defs, T["sheet_cmp"], cat), f"A{row + 1}")
        row += 21
    return ws


# ---------------- 统计汇总 ----------------
def station_stats(records):
    """records: [(qty, n1, n2, n3), ...] -> 统计字典。"""
    qty = sum(r[0] for r in records)
    noks = [sum(r[i] for r in records) for i in (1, 2, 3)]
    total_nok = sum(noks)
    daily = [(sum(r[1:]) / r[0]) if r[0] else 0 for r in records]
    return {
        "days": len(records), "qty": qty, "noks": noks, "total_nok": total_nok,
        "rates": [n / qty if qty else 0 for n in noks],
        "total_rate": total_nok / qty if qty else 0,
        "max_rate": max(daily), "max_idx": daily.index(max(daily)),
        "min_rate": min(daily), "min_idx": daily.index(min(daily)),
        "avg_qty": qty / len(records) if records else 0,
    }


def write_summary_sheet(wb, dates, data_a, data_b):
    ws = wb.create_sheet(T["sheet_sum"])
    ws["A1"] = T["sum_title"].format(d0=dates[0], d1=dates[-1], n=len(dates))
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 26

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])
    lb = T["sum_rows"]
    dash = "—"

    rows = [(T["sum_heads"], None)] + [
        ((lb[0], sa["days"], sb["days"], sb["days"] - sa["days"]), "int"),
        ((lb[1], sa["qty"], sb["qty"], sb["qty"] - sa["qty"]), "int"),
        ((lb[2], sa["avg_qty"], sb["avg_qty"], sb["avg_qty"] - sa["avg_qty"]), "f1"),
        ((lb[3], sa["noks"][0], sb["noks"][0], sb["noks"][0] - sa["noks"][0]), "int"),
        ((lb[4], sa["rates"][0], sb["rates"][0], sb["rates"][0] - sa["rates"][0]), "pct"),
        ((lb[5], sa["noks"][1], sb["noks"][1], sb["noks"][1] - sa["noks"][1]), "int"),
        ((lb[6], sa["rates"][1], sb["rates"][1], sb["rates"][1] - sa["rates"][1]), "pct"),
        ((lb[7], sa["noks"][2], sb["noks"][2], sb["noks"][2] - sa["noks"][2]), "int"),
        ((lb[8], sa["rates"][2], sb["rates"][2], sb["rates"][2] - sa["rates"][2]), "pct"),
        ((lb[9], sa["total_nok"], sb["total_nok"], sb["total_nok"] - sa["total_nok"]), "int"),
        ((lb[10], sa["total_rate"], sb["total_rate"], sb["total_rate"] - sa["total_rate"]), "pct"),
        ((lb[11], sa["max_rate"], sb["max_rate"], sb["max_rate"] - sa["max_rate"]), "pct"),
        ((lb[12], dates[sa["max_idx"]], dates[sb["max_idx"]], dash), "text"),
        ((lb[13], sa["min_rate"], sb["min_rate"], sb["min_rate"] - sa["min_rate"]), "pct"),
        ((lb[14], dates[sa["min_idx"]], dates[sb["min_idx"]], dash), "text"),
    ]

    for i, (values, kind) in enumerate(rows):
        r = 3 + i
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border, cell.alignment = BORDER, CENTER
            if i == 0:
                cell.font, cell.fill = FONT_HEAD, FILL_SUM
            else:
                cell.font = FONT_BODY if col == 1 else FONT_TOTAL
                if i % 2 == 0:
                    cell.fill = FILL_ALT
                if col > 1:
                    cell.number_format = {"pct": "0.00%", "f1": "0.0", "int": "#,##0"}.get(kind, "General")
        ws.cell(row=r, column=1).alignment = LEFT
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 30
    for col in "BCD":
        ws.column_dimensions[col].width = 18

    note_row = 3 + len(rows) + 1
    for i, text in enumerate(T["notes"]):
        cell = ws.cell(row=note_row + i, column=1,
                       value=text.format(chart=T["sheet_chart"], a=T["sheet_a"], b=T["sheet_b"],
                                         cmp=T["sheet_cmp"], ver=VERSION))
        cell.font = Font(name=T["font"], size=9, color="595959", bold=(i == 0))
        cell.alignment = LEFT


def write_raw_sheet(wb, src_path):
    """保留一份原始数据，便于回溯核对。"""
    src = load_workbook(src_path, data_only=True).worksheets[0]
    ws = wb.create_sheet(T["sheet_raw"])
    for row in src.iter_rows(min_row=3, max_row=src.max_row, max_col=9, values_only=True):
        if all(v is None for v in row):
            continue
        ws.append(list(row))
    if T["raw_heads"]:
        for col, head in enumerate(T["raw_heads"], start=1):
            ws.cell(row=1, column=col, value=head)
    for col in range(1, 10):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.alignment = FONT_HEAD, CENTER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="808080")
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


# ---------------- 主流程 ----------------
def main(src_path, out_path, lang):
    set_lang(lang)
    rows = read_raw(src_path)
    if not rows:
        print("no data rows found in the source workbook", file=sys.stderr)
        return 1

    dates = [r[0] for r in rows]
    data_a = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]
    data_b = [(r[0], r[5], r[6], r[7], r[8]) for r in rows]

    wb = Workbook()
    wb.remove(wb.active)
    ws_a = write_station_sheet(wb, T["sheet_a"], T["station_a"], data_a, FILL_A, RAMP_A)
    ws_b = write_station_sheet(wb, T["sheet_b"], T["station_b"], data_b, FILL_B, RAMP_B)
    write_compare_sheet(wb, data_a, data_b, dates)
    write_summary_sheet(wb, dates, data_a, data_b)
    write_raw_sheet(wb, src_path)
    chart_sheet = write_chart_sheet(wb, ws_a, ws_b, len(dates))
    wb.move_sheet(chart_sheet, offset=-(len(wb.sheetnames) - 1))  # 图表页放到第一位
    wb.active = 0
    wb.save(out_path)

    sa = station_stats([r[1:] for r in data_a])
    sb = station_stats([r[1:] for r in data_b])
    n_charts = sum(len(ws._charts) for ws in wb)
    print(f"written: {out_path}  (lang={lang}, {len(wb.sheetnames)} sheets, {n_charts} line charts)")
    print(f"{len(dates)} production days ({dates[0]} - {dates[-1]})")
    for name, s in (("A", sa), ("B", sb)):
        print(f"Station {name}: output {s['qty']}, NOK {s['total_nok']}, total {s['total_rate']:.2%}, "
              f"PDIV1/2/3 = {s['rates'][0]:.2%} / {s['rates'][1]:.2%} / {s['rates'][2]:.2%}")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    lang = "en" if "--lang=en" in sys.argv or ("--lang" in sys.argv and "en" in sys.argv) else "zh"
    for a in sys.argv[1:]:
        if a.startswith("--lang="):
            lang = a.split("=", 1)[1]
    src = Path(args[0]) if args else Path("New_Microsoft_Excel_Worksheet.xlsx")
    default = f"PDIV_Inspection_Statistics_A_vs_B_v1.2.xlsx" if lang == "en" \
        else "PDIV检测数据统计_A站B站_v1.2.xlsx"
    out = Path(args[1]) if len(args) > 1 else Path(default)
    sys.exit(main(src, out, lang))
